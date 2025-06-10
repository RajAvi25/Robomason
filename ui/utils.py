## ui/utils.py

import os
import re
import subprocess
from pathlib import Path
import pickle
import shutil

import base64

import threading
import time

import pandas as pd

from matplotlib import pyplot as plt
from IPython.display import display
from matplotlib.figure import Figure

import ui
from construct import analysis
import camera

from ui.explanations import explanations 

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import webbrowser

def show_project_structure():
    directory = "/home/avi/Desktop/robomason"
    for root, _, files in os.walk(directory):
        level = root.replace(directory, "").count(os.sep)
        indent_space = " " * 4 * level
        print(f"{indent_space}- {os.path.basename(root)}/")  # Print folder name
        for file in files:
            print(f"{indent_space}    - {file}") 

def initializeUI():
    ui.system_setup()
    framehandler = camera.FrameHandler(ws_url="ws://localhost:9090", camera_index=4, frame_rate=15, is_sender=False)
    threading.Thread(target=framehandler.start_streaming, daemon=True).start()
    time.sleep(1.5)
    frame = framehandler.get_latest_frame()
    plt.imshow(frame)
    return framehandler


def analyze_trajectory(options,construction_type):
    """
    Performs various types of analyses and plotting on trajectory and detection data 
    based on the specified options.

    Args:
        options (dict): A dictionary containing flags to enable different types of analyses,
                        and the input data required for processing.
            Expected keys:
                - 'data' (DataFrame): The input dataset containing trajectory and detection information.
                - 'trajectory_analysis' (bool): If True, performs trajectory metric analysis and plots kinematic data.
                - 'hazard_analysis' (bool): If True, processes and visualizes hazard events (used with trajectory_analysis).
                - 'detection_analysis' (bool): If True, performs worker detection analysis.
                - 'plot_full_run' (bool): If True, plots the full trajectory run.
                - 'segment_analysis' (bool): If True, generates segment-wise analysis plots.
                - 'plot_workers' (bool): If True, plots worker positions on the site.

    Returns:
        dict: A dictionary containing results based on the selected analyses, including:
            - 'hazard_summary': Summary table of detected hazard events (if applicable).
            - 'hazard_plot': Plot of hazard event timeline (if applicable).
            - 'trajectory_summary': Metrics and statistics for the full trajectory.
            - 'Complete_kinematic_plots': List containing velocity and acceleration plots.
            - 'worker_detection': Results of worker detection analysis.
            - 'Trajectory plots': Visualization of the complete trajectory.
            - 'segment_plots': Plots from segment-wise analysis.
            - 'worker plot': Visualization of detected worker positions.
    """

    data = options.get("data", None)
    results = {}

    if options.get("trajectory_analysis", False):
        print("[INFO] Starting trajectory analysis...")
        evnt = None
        if options.get("hazard_analysis", False):
            print("[INFO] Performing hazard event processing...")
            _hazard_events = analysis.process_hazard_events(data)
            print("[INFO] Tabulating hazard events...")
            results["hazard_summary"] = analysis.tabulate_hazard_events(_hazard_events)
            print("[INFO] Plotting hazard events...")
            results["hazard_plot"] = analysis.plot_hazard_events(_hazard_events)
            evnt = _hazard_events
        print("[INFO] Summarizing trajectory metrics...")
        results["trajectory_summary"] = analysis.summarize_trajectory_metrics(data, construction_type="assembly", hazard_events=evnt)
        print("[INFO] Plotting velocity and acceleration...")
        fig_vel, fig_acc, _, _, _ = analysis.complete_velocity_acceleration_plot(data)
        results["Complete_kinematic_plots"] = [fig_vel, fig_acc]
        print("[INFO] Trajectory analysis completed.")

    if options.get("detection_analysis", False):
        print("[INFO] Starting worker detection analysis...")
        results["worker_detection"] = analysis.workerdetection_analysis(data)
        print("[INFO] Worker detection analysis completed.")

    if options.get("segment_analysis", False):
        print("[INFO] Performing segment-wise analysis...")
        results["segment_plots"] = analysis.analyse_segment(data)
        print("[INFO] Segment analysis completed.")

    if options.get("plot_workers", False):
        print("[INFO] Plotting worker positions...")
        results["worker plot"] = analysis.plot_workers(data,False) # This bool enables a 3D worker detection plot
        print("[INFO] Worker plotting completed.")

    if options.get("plot_full_run", False):
        print("[INFO] Plotting full trajectory run...")
        results["Trajectory plots"] = analysis.plot_complete_data(data, 'Complete')
        print("[INFO] Full run plot completed.")

    print("[INFO] All requested analyses completed.")
    return results


def load_data(n):
    # Retrieve current tracking data.
    data = ui.comms.get_tracking_packets()
    
    # Directory to store logs.
    directory = '/home/avi/Desktop/robomason/_workingdata/_rawtrajectories'
    
    # List files that are pickle files (assuming names like "0.pkl", "1.pkl", etc.)
    files = [f for f in os.listdir(directory) if f.endswith('.pkl')]
    
    # Extract numeric parts from file names.
    log_numbers = []
    for filename in files:
        try:
            # Remove the file extension and convert to an integer.
            log_num = int(os.path.splitext(filename)[0])
            log_numbers.append(log_num)
        except ValueError:
            # If filename does not match our numeric pattern, skip it.
            continue
    
    # Decide on the file name based on the count of existing log files.
    if len(log_numbers) < n:
        # Determine next log number.
        next_log_num = max(log_numbers) + 1 if log_numbers else 0
        file_path = os.path.join(directory, f"{next_log_num}.pkl")
    else:
        # If there are already n logs, delete all existing log files.
        for f in files:
            os.remove(os.path.join(directory, f))
        # Start fresh with file name 0.
        file_path = os.path.join(directory, "0.pkl")
    
    # Save the data to a pickle file.
    with open(file_path, "wb") as file:
        pickle.dump(data, file)
    
    return data

def save_trajectory_summary_excel(df, file_path):
    """
    Save the trajectory summary DataFrame to an Excel file with improved styling using openpyxl.
    
    Args:
        df (pd.DataFrame): The trajectory summary DataFrame.
        file_path (str): Full path (including .xlsx) where the file will be saved.
    """
    # Create a new workbook and select the active worksheet.
    wb = Workbook()
    ws = wb.active
    ws.title = "Trajectory Summary"

    # Define some styles.
    header_font = Font(name='Calibri', bold=True, color="000000")  # black text
    header_fill = PatternFill(start_color="C9C9C9", end_color="C9C9C9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1) Write column headers with styling.
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # 2) Write DataFrame rows, applying borders/alignment.
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            # Convert lists/tuples to string to avoid ValueError
            if isinstance(value, (list, tuple)):
                value = ", ".join(str(x) for x in value)

            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border

    # 3) Auto-fit column widths based on max cell length in each column.
    #    Note: openpyxl does not have a built-in auto-fit; we approximate by measuring text length.
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            cell_value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(cell_value))
        # A little extra spacing
        ws.column_dimensions[column_letter].width = max_length + 2

    # 4) Freeze the top row so headers remain visible when scrolling.
    ws.freeze_panes = "A2"

    # 5) Save the workbook.
    wb.save(file_path)

def save_worker_detection_excel(df, file_path):
    """
    Save the worker detection DataFrame to an Excel file with improved styling using openpyxl.
    - The 'first detected at location (time)' column is modified to show only the time (no coordinates).
    - The header row height is increased to accommodate wrapped text.
    
    Args:
        df (pd.DataFrame): The worker detection DataFrame.
        file_path (str): Full path (including .xlsx) where the file will be saved.
    """
    # Create a new workbook and select the active worksheet.
    wb = Workbook()
    ws = wb.active
    ws.title = "Worker Detection"

    # Define some styles.
    header_font = Font(name='Calibri', bold=True, color="000000")  # black text
    header_fill = PatternFill(start_color="C9C9C9", end_color="C9C9C9", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 1) Write column headers with styling.
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Increase the header row height so wrapped text is fully visible.
    ws.row_dimensions[1].height = 30

    # 2) Write DataFrame rows, applying borders/alignment.
    for row_idx, row in df.iterrows():
        for col_idx, value in enumerate(row, start=1):
            # Convert lists/tuples to string to avoid ValueError
            if isinstance(value, (list, tuple)):
                value = ", ".join(str(x) for x in value)

            # If this is the "first detected at location (time)" column, strip out coordinates, keep only time
            col_header = df.columns[col_idx - 1]
            if col_header == "first detected at location (time)" and isinstance(value, str):
                # Look for text inside parentheses (time) and replace the cell value with just that
                match = re.search(r"\((.*?)\)", value)
                if match:
                    value = match.group(1)  # only the time portion

            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
            cell.alignment = center_alignment
            cell.border = thin_border

    # 3) Auto-fit column widths based on max cell length in each column.
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            cell_value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = max_length + 2

    # 4) Freeze the top row so headers remain visible when scrolling.
    ws.freeze_panes = "A2"

    # 5) Save the workbook.
    wb.save(file_path)


def save_results(results, data, mode='ct'):
    """
    Saves analysis results and raw data to a new subfolder.
    
    The function creates a new subfolder inside
    /home/avi/Desktop/robomason/_workingdata/{subfolder}
    with a name of the form '{run_type}_run_XX', where XX is the next available run number.
    
    DataFrames in results are saved as Excel files.
    Figures (from Matplotlib) are saved as high-quality SVG images.
    The raw data is saved as a pickle file.
    
    Args:
        results (dict): A dictionary with keys:
            - "trajectory_summary": DataFrame to be saved as Excel.
            - "worker_detection": DataFrame to be saved as Excel.
            - "Trajectory plots": List of figures to be saved as SVGs.
            - "segment_plots": List of dictionaries (one per segment). Each dictionary is expected to have:
                   * "segment": a string representing the segment name.
                   * "kinematic plots": a list of figures.
                   * "trajectory_plots": either a dictionary (keys = state names, values = list of figures)
                                          or a list of figures.
                   * Optionally, "FOV": a list of figures.
        data: Raw data to be saved as a pickle file.
        mode (str): Determines the subfolder type. Options:
            - 'ct' (default) -> "_constructionruns"
            - 'dis' -> "_disassemblyruns"
            - 'rect' -> "_reconstructionruns"
            - 'full' -> "_fullruns"
    """
    # Determine the base directory and run type based on mode
    mode_mapping = {
        'ct': ('_constructionruns', 'construction'),
        'dis': ('_disassemblyruns', 'disassembly'),
        'rect': ('_reconstructionruns', 'reconstruction'),
        'full': ('_fullruns', 'complete_construction')
    }
    
    if mode not in mode_mapping:
        raise ValueError("Invalid mode. Choose from 'ct', 'dis', 'rect', or 'full'.")
    
    subfolder, run_type = mode_mapping[mode]
    base_dir = f'/home/avi/Desktop/robomason/_workingdata/{subfolder}'
    os.makedirs(base_dir, exist_ok=True)

    # List all subdirectories matching our naming scheme.
    subfolders = [name for name in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, name))]
    run_numbers = []
    pattern = re.compile(fr'{run_type}_run_(\d+)')
    for folder in subfolders:
        match = pattern.match(folder)
        if match:
            run_numbers.append(int(match.group(1)))
    next_run = max(run_numbers) + 1 if run_numbers else 0
    new_folder_name = f"{run_type}_run_{next_run:02d}"
    new_folder_path = os.path.join(base_dir, new_folder_name)
    os.makedirs(new_folder_path, exist_ok=True)
    
    # Save DataFrames
    if "trajectory_summary" in results:
        traj_summary_path = os.path.join(new_folder_path, "trajectory_summary.xlsx")
        save_trajectory_summary_excel(results["trajectory_summary"], traj_summary_path)
    
    if "Complete_kinematic_plots" in results:
        fig_vel, fig_acc = results["Complete_kinematic_plots"]
        vel_path = os.path.join(new_folder_path, "complete_velocity_plot.svg")
        acc_path = os.path.join(new_folder_path, "complete_acceleration_plot.svg")
        if isinstance(fig_vel, Figure):
            fig_vel.savefig(vel_path, format="svg")
        if isinstance(fig_acc, Figure):
            fig_acc.savefig(acc_path, format="svg")
    
    if "hazard_summary" in results:
        results["hazard_summary"].to_excel(os.path.join(new_folder_path, "hazard_summary.xlsx"), index=False)
    
    if "hazard_plot" in results:
        results["hazard_plot"].savefig(os.path.join(new_folder_path, "marker detection events.svg"), format="svg")
    
    if "worker_detection" in results:
        worker_detect_path = os.path.join(new_folder_path, "worker_detection.xlsx")
        save_worker_detection_excel(results["worker_detection"], worker_detect_path)
    
    # Save global trajectory plots
    if "Trajectory plots" in results:
        figures = results["Trajectory plots"]
        view_names = ["Isometric view", "Plan view", "Side view", "Front view"]
        for idx, fig in enumerate(figures):
            if idx < len(view_names):
                fig_path = os.path.join(new_folder_path, f"trajectory_{view_names[idx]}.svg")
                fig.savefig(fig_path, format="svg")
    
    # Save segment-specific plots
    if "segment_plots" in results:
        for segment_dict in results["segment_plots"]:
            # Convert segment identifier to string, then clean it.
            raw_segment = str(segment_dict.get("segment", "Unnamed"))
            clean_segment = raw_segment.strip().replace("/", "-")
            segment_folder = os.path.join(new_folder_path, clean_segment)
            os.makedirs(segment_folder, exist_ok=True)
    
            # Save kinematic plots
            for i, fig in enumerate(segment_dict.get('kinematic plots', [])):
                if isinstance(fig, Figure):
                    fig.savefig(os.path.join(segment_folder, f"plot_{clean_segment}_kinematic_{i}.svg"), format="svg")
    
            # Save trajectory plots
            traj_plots = segment_dict.get("trajectory_plots")
            if traj_plots:
                if isinstance(traj_plots, dict):
                    # For each state key (like "search", "pick", etc.), save its list of figures.
                    for state, figs in traj_plots.items():
                        state_folder = os.path.join(segment_folder, state)
                        os.makedirs(state_folder, exist_ok=True)
                        # If figs is not a list, wrap it.
                        if not isinstance(figs, list):
                            figs = [figs]
                        for i, fig in enumerate(figs):
                            if isinstance(fig, Figure):
                                fig.savefig(os.path.join(state_folder, f"trajectory_{clean_segment}_{state}_{i}.svg"), format="svg")
                elif isinstance(traj_plots, list):
                    # Otherwise, if trajectory_plots is just a list, save them all.
                    for i, fig in enumerate(traj_plots):
                        if isinstance(fig, Figure):
                            fig.savefig(os.path.join(segment_folder, f"trajectory_{clean_segment}_{i}.svg"), format="svg")
    
            # Save FOV plots, if any.
            for i, fig in enumerate(segment_dict.get("FOV", [])):
                if isinstance(fig, Figure):
                    fig.savefig(os.path.join(segment_folder, f"FOV_plot_{i}.svg"), format="svg")
    
    if "worker plot" in results:
        figures = results["worker plot"]
        view_names = ["Isometric view", "Plan view"]
        for idx, fig in enumerate(figures):
            if fig is None:
                continue
            if idx < len(view_names):
                fig_path = os.path.join(new_folder_path, f"worker_{view_names[idx]}.svg")
                fig.savefig(fig_path, format="svg")
    
    # Save the raw data as a pickle file.
    raw_data_path = os.path.join(new_folder_path, "raw_data.pkl")
    with open(raw_data_path, "wb") as file:
        pickle.dump(data, file)
    
    print(f"Results saved in {new_folder_path}")
    return new_folder_path

def open_html_in_browser(html_file_path):
    """
    Opens the specified HTML file in the default web browser.
    
    Parameters:
        html_file_path (str): The full path to the HTML file.
        save_path (str): The destination path where the HTML should be saved.
    """
    webbrowser.open(f"file:///{html_file_path}")


def generate_segment(base_path):
    """
    Generates HTML for the segment analysis by iterating over each segment folder
    in the base_path. For each segment, images are embedded in a defined order for 
    the trajectory plots: search, pick, swing, place, swing_back. If a given state 
    folder does not exist (as in disassembly runs which lack "search"), it is skipped.
    
    Parameters:
        base_path (str): The base folder path containing the segment subfolders.
    
    Returns:
        str: HTML snippet with segment analysis.
    """
    # List all segment subfolders in the base_path
    segment_folders = [folder for folder in os.listdir(base_path)
                       if os.path.isdir(os.path.join(base_path, folder))]
    
    # Optional mapping: convert "Toilet_1" etc. into "Bathroom Module 1"
    toilet_to_bathroom = {
        "Toilet_1": "Bathroom Module 1",
        "Toilet_2": "Bathroom Module 2",
        "Toilet_3": "Bathroom Module 3"
    }
    
    # Define the desired state order. If a state folder is absent, it will be skipped.
    ordered_state_names = ["search", "pick", "swing", "place", "swing_back"]
    
    html = ""
    for folder in sorted(segment_folders):
        folder_path = os.path.join(base_path, folder)
        # Use mapping if available
        display_name = toilet_to_bathroom.get(folder, folder)
        html += f"<div class='section'><h3>{display_name}</h3>"
    
        # --- Kinematic Plots ---
        html += "<div class='image-grid'>"
        # Here we assume kinematic plots are named as 'plot_{folder}_kinematic_{i}.svg'.
        # Change the range if you expect more plots.
        for i in range(2):
            kin_plot = os.path.join(folder_path, f"plot_{folder}_kinematic_{i}.svg")
            if os.path.isfile(kin_plot):
                data_uri = encode_image_to_data_uri(kin_plot, mime_type="image/svg+xml")
                if data_uri:
                    html += f'<img src="{data_uri}" alt="Kinematic Plot {i}" style="width:90%; max-width:900px; border:1px solid #ccc;">'
        html += "</div>"
    
        # --- Trajectory Plots (ordered) ---
        # For each state in the ordered list, check if a folder exists and then process it.
        for state in ordered_state_names:
            state_folder = os.path.join(folder_path, state)
            if not os.path.isdir(state_folder):
                continue  # Skip missing state folder (e.g. "search" in disassembly)
            
            # Add a header for this state.
            html += f"<h4 style='margin-top: 20px;'>{state.capitalize()}</h4><div class='image-grid'>"
            files = sorted(os.listdir(state_folder))
            for f in files:
                if f.endswith(".svg"):
                    file_path = os.path.join(state_folder, f)
                    data_uri = encode_image_to_data_uri(file_path, mime_type="image/svg+xml")
                    if data_uri:
                        html += (f'<img src="{data_uri}" alt="{f}" style="width:30%; min-width:280px; max-width:360px; border:1px solid #ccc;">')
            html += "</div>"
        html += "</div>"
    return html

def encode_image_to_data_uri(filepath, mime_type="image/svg+xml"):
    """
    Reads the image file from filepath, encodes it in base64, and returns a data URI.
    
    Parameters:
        filepath (str): Path to the image file.
        mime_type (str): MIME type, default is "image/svg+xml".
    
    Returns:
        str: Data URI string for embedding.
    """
    try:
        with open(filepath, "rb") as f:
            data = f.read()
        encoded = base64.b64encode(data).decode("utf-8")
        return f"data:{mime_type};base64,{encoded}"
    except Exception as e:
        print(f"Error encoding {filepath}: {e}")
        return ""

def create_construction_summary_html(segment_folder_path,
                                     table_max_height: int = 400):
    """
    Generates an HTML summary page in `segment_folder_path/Accumulated data.html`.
    Copies all detected SVGs into an `images/` subfolder and references them
    with purely relative paths so you can share the folder intact.

    Parameters:
        segment_folder_path (str): folder containing .xlsx and .svg files
        table_max_height (int): max-px height for scrollable tables

    Returns:
        str: absolute path to the generated HTML file
    """
    # --- Setup output paths ---
    output_file   = os.path.join(segment_folder_path, "Accumulated data.html")
    images_dir    = os.path.join(segment_folder_path, "images")
    os.makedirs(images_dir, exist_ok=True)
    images_folder = "images"  # relative folder name for src attributes

    # --- Title from folder structure ---
    p1 = os.path.dirname(segment_folder_path)
    p2 = os.path.dirname(p1)
    diff = os.path.relpath(p1, p2).replace("_", " ").replace("runs", " Run").title().strip()
    html_title = f"{diff} Summary"

    # --- Helpers ---
    def img_tag(src_path, alt_text):
        # copy into images/ if not already there
        fname = os.path.basename(src_path)
        dst   = os.path.join(images_dir, fname)
        if not os.path.exists(dst):
            shutil.copy2(src_path, dst)
        # always use relative path "images/<fname>"
        rel = f"{images_folder}/{fname}"
        return f'<img src="{rel}" alt="{alt_text}" loading="lazy">'

    def wrap_table(df):
        html = df.to_html(index=False, classes="dataframe", border=0)
        return (f'<div style="max-height:{table_max_height}px; '
                f'overflow:auto; margin-bottom:20px;">{html}</div>')

    # --- Data paths ---
    conf_xlsx   = os.path.join(segment_folder_path, "trajectory_summary.xlsx")
    worker_xlsx = os.path.join(segment_folder_path, "worker_detection.xlsx")
    hazard_xlsx = os.path.join(segment_folder_path, "hazard_summary.xlsx")
    marker_svg  = os.path.join(segment_folder_path, "marker detection events.svg")

    # image filenames (in segment_folder_path)
    worker_svgs = ["worker_Isometric view_.svg", "worker_Plan view.svg"]
    traj_svgs   = [
        "trajectory_Isometric view.svg",
        "trajectory_Plan view.svg",
        "trajectory_Side view.svg",
        "trajectory_Front view.svg",
        "complete_velocity_plot.svg",
        "complete_acceleration_plot.svg"
    ]

    sections = []

    # 1) Kinematics & Trajectory
    traj_paths = [os.path.join(segment_folder_path, f) for f in traj_svgs]
    avail = [p for p in traj_paths if os.path.exists(p)]
    if avail:
        html = ['''
    <div class="section">
      <h2>Kinematics & Trajectory</h2>
      <div class="image-grid">
        ''']
        for p in avail:
            alt = os.path.splitext(os.path.basename(p))[0].replace("_", " ").capitalize()
            html.append(img_tag(p, alt))
        html.append('''
      </div>
    </div>''')
        sections.append("".join(html))

    # 2) Construction Summary
    if os.path.exists(conf_xlsx):
        try:
            df = pd.read_excel(conf_xlsx)
            tbl = wrap_table(df)
            key = os.path.relpath(p1, p2)
            title_map = {
                "_constructionruns": "Assembly Summary",
                "_disassemblyruns":  "Disassembly Summary",
                "_reconstructionruns":"Reassembly Summary"
            }
            sec_title = title_map.get(key, "Construction Summary")
            html = [f'''
    <div class="section">
      <h2>{sec_title}</h2>
      {tbl}
      <div class="explanation">
            ''']
            for k, txt in explanations.items():
                if "Formula" in k:
                    html.append(f'<p class="formula">{txt}</p>')
                else:
                    html.append(f'<p><strong>{k}</strong>: {txt}</p>')
            html.append('''
      </div>
    </div>''')
            sections.append("".join(html))
        except Exception as e:
            print("Construction summary error:", e)

    # 3) Hazard Analysis
    if os.path.exists(hazard_xlsx) and os.path.exists(marker_svg):
        try:
            dfh = pd.read_excel(hazard_xlsx)
            tbl_h = wrap_table(dfh)
            img_h = img_tag(marker_svg, "Marker Detection Events")
            html = f'''
    <div class="section">
      <h2>Hazard Analysis</h2>
      {tbl_h}
      <div class="image-grid" style="margin-top:20px;">
        {img_h}
      </div>
    </div>'''
            sections.append(html)
        except Exception as e:
            print("Hazard analysis error:", e)

    # 4) Workers & Work Zones
    wpaths = [os.path.join(segment_folder_path, f) for f in worker_svgs]
    wavail = [p for p in wpaths if os.path.exists(p)]
    if wavail:
        html = ['''
    <div class="section">
      <h2>Workers and Work Zones</h2>
      <div class="image-grid">
        ''']
        for p in wavail:
            html.append(img_tag(p, "Worker View"))
        html.append('''
      </div>
    </div>''')
        sections.append("".join(html))

    # 5) Worker Detection
    if os.path.exists(worker_xlsx):
        try:
            dfw = pd.read_excel(worker_xlsx)
            tbl_w = ("<p>No worker detection data available.</p>"
                     if dfw.empty else wrap_table(dfw))
            html = f'''
    <div class="section">
      <h2>Worker Detection</h2>
      {tbl_w}
    </div>'''
            sections.append(html)
        except Exception as e:
            print("Worker detection error:", e)

    # 6) Segment Analysis
    seg_html = generate_segment(segment_folder_path)
    if seg_html.strip():
        sections.append(f'''
    <div class="section">
      <h2>Segment Analysis</h2>
      {seg_html}
    </div>''')

    # --- Write HTML ---
    header = f'''<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>{html_title}</title>
  <style>
    /* your CSS here… */
  </style>
  <script async
    src="https://cdnjs.cloudflare.com/ajax/libs/mathjax/2.7.7/MathJax.js?config=TeX-MML-AM_CHTML">
  </script>
</head>
<body>
  <h1>{html_title}</h1>
'''
    footer = '''
</body>
</html>'''

    with open(output_file, "w", encoding="utf-8") as f:
        f.write(header)
        for sec in sections:
            f.write(sec + "\n")
        f.write(footer)

    return os.path.abspath(output_file)

def save_results(results, data, mode):
    """
    Saves analysis results and raw data to a new subfolder.
    
    The function creates a new subfolder inside
    /home/avi/Desktop/robomason/_workingdata/{subfolder}
    with a name of the form '{run_type}_run_XX', where XX is the next available run number.
    
    DataFrames in results are saved as Excel files.
    Figures (from Matplotlib) are saved as high-quality SVG images.
    The raw data is saved as a pickle file.
    
    Args:
        results (dict): A dictionary with keys:
            - "trajectory_summary": DataFrame to be saved as Excel.
            - "worker_detection": DataFrame to be saved as Excel.
            - "Trajectory plots": List of figures to be saved as SVGs.
            - "segment_plots": List of dictionaries representing segment results.
        data: Raw data to be saved as a pickle file.
        mode (str): Determines the subfolder type. Options:
            - 'ct' (default) -> "_constructionruns"
            - 'dis' -> "_disassemblyruns"
            - 'rect' -> "_reconstructionruns"
            - 'full' -> "_fullruns"
    """
    # Determine the base directory and run type based on mode
    mode_mapping = {
        'ct': ('_constructionruns', 'construction'),
        'dis': ('_disassemblyruns', 'disassembly'),
        'rect': ('_reconstructionruns', 'reconstruction'),
        'full': ('_fullruns', 'complete_construction')
    }
    
    if mode not in mode_mapping:
        raise ValueError("Invalid mode. Choose from 'ct', 'dis', 'rect', or 'full'.")
    
    subfolder, run_type = mode_mapping[mode]
    base_dir = f'/home/avi/Desktop/robomason/_workingdata/{subfolder}'
    os.makedirs(base_dir, exist_ok=True)

    subfolders = [name for name in os.listdir(base_dir) if os.path.isdir(os.path.join(base_dir, name))]
    run_numbers = []
    pattern = re.compile(fr'{run_type}_run_(\d+)')
    for folder in subfolders:
        match = pattern.match(folder)
        if match:
            run_numbers.append(int(match.group(1)))
    next_run = max(run_numbers) + 1 if run_numbers else 0
    new_folder_name = f"{run_type}_run_{next_run:02d}"
    new_folder_path = os.path.join(base_dir, new_folder_name)
    os.makedirs(new_folder_path, exist_ok=True)
    
    # Save DataFrames
    if "trajectory_summary" in results:
        traj_summary_path = os.path.join(new_folder_path, "trajectory_summary.xlsx")
        save_trajectory_summary_excel(results["trajectory_summary"], traj_summary_path)
    
    if "Complete_kinematic_plots" in results:
        fig_vel, fig_acc = results["Complete_kinematic_plots"]
        vel_path = os.path.join(new_folder_path, "complete_velocity_plot.svg")
        acc_path = os.path.join(new_folder_path, "complete_acceleration_plot.svg")
        if isinstance(fig_vel, Figure):
            fig_vel.savefig(vel_path, format="svg")
        if isinstance(fig_acc, Figure):
            fig_acc.savefig(acc_path, format="svg")
    
    if "hazard_summary" in results:
        results["hazard_summary"].to_excel(os.path.join(new_folder_path, "hazard_summary.xlsx"), index=False)
    
    if "hazard_plot" in results:
        results["hazard_plot"].savefig(os.path.join(new_folder_path, "marker detection events.svg"), format="svg")
    
    if "worker_detection" in results:
        worker_detect_path = os.path.join(new_folder_path, "worker_detection.xlsx")
        save_worker_detection_excel(results["worker_detection"], worker_detect_path)
    
    if "Trajectory plots" in results:
        figures = results["Trajectory plots"]
        view_names = ["Isometric view", "Plan view", "Side view", "Front view"]
        for idx, fig in enumerate(figures):
            if idx < len(view_names):
                fig_path = os.path.join(new_folder_path, f"trajectory_{view_names[idx]}.svg")
                fig.savefig(fig_path, format="svg")
    
    # Updated segment-specific plots saving logic
    if "segment_plots" in results:
        for segment_dict in results["segment_plots"]:
            raw_segment = str(segment_dict.get("segment", "Unnamed"))
            clean_segment = raw_segment.strip().replace("/", "-")
            segment_folder = os.path.join(new_folder_path, clean_segment)
            os.makedirs(segment_folder, exist_ok=True)
    
            # Save kinematic plots
            for i, fig in enumerate(segment_dict.get('kinematic plots', [])):
                if isinstance(fig, Figure):
                    fig.savefig(os.path.join(segment_folder, f"plot_{clean_segment}_kinematic_{i}.svg"), format="svg")
    
            # Save trajectory plots:
            traj_plots = segment_dict.get("trajectory_plots")
            if traj_plots:
                if isinstance(traj_plots, dict):
                    # Iterate through each state key, which could be "search", "pick", etc.
                    for state_key, figs in traj_plots.items():
                        state_folder = os.path.join(segment_folder, state_key)
                        os.makedirs(state_folder, exist_ok=True)
                        if not isinstance(figs, list):
                            figs = [figs]
                        for i, fig in enumerate(figs):
                            if isinstance(fig, Figure):
                                fig.savefig(os.path.join(state_folder, f"trajectory_{clean_segment}_{state_key}_{i}.svg"), format="svg")
                elif isinstance(traj_plots, list):
                    # Otherwise, if it's a simple list, save all in the segment folder.
                    for i, fig in enumerate(traj_plots):
                        if isinstance(fig, Figure):
                            fig.savefig(os.path.join(segment_folder, f"trajectory_{clean_segment}_{i}.svg"), format="svg")
    
            # Save FOV plots, if any.
            for i, fig in enumerate(segment_dict.get("FOV", [])):
                if isinstance(fig, Figure):
                    fig.savefig(os.path.join(segment_folder, f"FOV_plot_{i}.svg"), format="svg")
    
    if "worker plot" in results:
        figures = results["worker plot"]
        view_names = ["Isometric view", "Plan view"]
        for idx, fig in enumerate(figures):
            if fig is None:
                continue
            if idx < len(view_names):
                fig_path = os.path.join(new_folder_path, f"worker_{view_names[idx]}.svg")
                fig.savefig(fig_path, format="svg")
    
    # Save raw data as pickle.
    raw_data_path = os.path.join(new_folder_path, "raw_data.pkl")
    with open(raw_data_path, "wb") as file:
        pickle.dump(data, file)
    
    print(f"Results saved in {new_folder_path}")
    return new_folder_path

